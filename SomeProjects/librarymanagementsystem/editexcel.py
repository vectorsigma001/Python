from models import *
import openpyxl

path = "students.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# this code will add the student
def addstudent(name,address,phonenumber):    
    for i in range(1, 501):  
        if sheet_obj.cell(row=i, column=1).value is None:
            c1 = sheet_obj.cell(row=i, column=1)
            c1.value = name
            c1 = sheet_obj.cell(row=i,column=2)
            c1.value = address
            c1 = sheet_obj.cell(row=i,column=3)
            c1.value = phonenumber
            break  
    wb_obj.save(path)

#path for books excel file
book_path = "bookdesc.xlsx"
book_obj = openpyxl.load_workbook(book_path)
book_sheet = book_obj.active

# this code will add the book
def addbook(name,description,author,stock):
    for i in range(1,501):
        if book_sheet.cell(row=i, column=1).value is None:
            b1 = book_sheet.cell(row=i,column=1)
            b1.value = name
            b1 = book_sheet.cell(row=i,column=2)
            b1.value = description
            b1 = book_sheet.cell(row=i,column=3)
            b1.value = author
            b1 = book_sheet.cell(row=i,column=4)
            b1.value = stock
            break
    book_obj.save(book_path)
    print("Product has been added successfully")


"""
give me the phonenumber and I will give you the information of that row
"""
phonenumberlist = []
for i in range(2, 501):
    cell_obj = sheet_obj.cell(row=i, column=3)
    if cell_obj.value is not None:
        phonenumberlist.append(cell_obj.value)
print(phonenumberlist)

def search_phonenumber(phonenumber):
    for i in range(2, 501):
        cell_obj = sheet_obj.cell(row=i, column=3)
        if cell_obj.value == phonenumber:
            return i  # Return the row number if the phone number is found
    return None 
studentrow = search_phonenumber(111)
print(search_phonenumber(111))


student_max_column = sheet_obj.max_column
for i in range(1, student_max_column + 1):
    cell_obj = sheet_obj.cell(row = studentrow, column=i)
    print(cell_obj.value)


# give me the author and name and I will give you the row information
# storing all the author row in a list 

"""
give me the author and name of the book and I will give you the book row data
"""

def search_authors(authorname, bookname):
    for row in range(2, book_sheet.max_row + 1):  
        current_author = book_sheet.cell(row=row, column=3).value
        current_book = book_sheet.cell(row=row, column=1).value
        if current_author == authorname and current_book == bookname:
            return row
    return None

# Function to get all information from a specific row
def get_row_information(row):
    if row:
        row_data = []
        for col in range(1, book_sheet.max_column + 1):
            cell_value = book_sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        return row_data
    else:
        return None
    
    
authorname = input("Enter the author name: ")
bookname = input("Enter the book name: ")
book_row = search_authors(authorname, bookname)

if book_row:
    print(f"Book '{bookname}' by '{authorname}' found at row {book_row}.")
    row_information = get_row_information(book_row)
    print("Row information:", row_information)
else:
    print(f"Book '{bookname}' by '{authorname}' not found.")
    

